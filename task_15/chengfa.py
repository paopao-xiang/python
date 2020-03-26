# -*- coding: utf-8 -*-
# @Time    : 2019/2/28 22:28
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : chengfa.py

# 5：输出99乘法表
# 1*1=1 i=1 j=1 range(1,i+1)
# 1*2=2 2*2=4 i=2 j=1 2 range(1,i+1)
# 1*3=3 2*3=6  3*3=9 i=3 j=1 2 3 range(1,i+1)
# 1*4=4 2*4=8  3*4=12 4*4=16 i=4 j=1 2 3 4  range(1,i+1)
# 1*5=5 2*5=10 3*5=15 4*5=20  5*5=25 i=5  j=1 2 3 4 5 range(1,i+1)
# for a in range(1,10):
#     for b in range(1,10):
#         print(a,"*",b,"=",a*b)
#1、一个足球队在寻找年龄在10岁到12岁的小女孩（包括10岁和12岁）加入。
# 编写一个程序，询问用户的性别（m表示男性，f表示女性）和年龄，
# 然后显示一条消息指出这个人是否可以加入球队，
# 询问10次后，输出满足条件的总人数。
#年龄 10<=age<=12  性别：女生  m  f
#询问10个人  问的问题是一样的~~ 我们就可以循环 for循环
#输出符合条件的总人数
# a=0
# b=0
# while a<10:
#     a+=1
#     age=input("age:")
#     six=input("sex:")
#     if 10<=int(age)<=12 and six=="m":
#         b+=1
# print(b)

# 生成随机整数，从1-9取出来。然后输入一个数字，来猜，
# 如果大于，则打印bigger。
# 小了，则打印less。如果相等，则打印equal
# import random
# num=random.randint(1,9)
# num1=int(input())
# if num>num1:
#     print("bigger")
# elif num<num1:
#     print("less")
# else:
#     print("equal")


#一家商场在降价促销。
# 如果购买金额50-100元(包含50元和100元)之间，会给10%的折扣，
# 如果购买金额大于100元会给20%折扣。
# 编写一程序，询问购买价格，再显示出折扣（%10或20%）和最终价格
#input() 从控制台获取商品总价格
# a=int(input("请问这件衣服多少钱："))
# if 50<=a<=100:
#     print("你好这件衣服",a,"元，所以可以打9折，最终价格是：",a*0.9)
# elif a>100:
#     print("你好这件衣服",a,"元，所以可以打9折，最终价格是：",a*0.8)
# else:
#     print("你好这件衣服",a,"元")

# 用嵌套for循环输出直角三角形
# *
# * *
# * * *
# * * * *
# * * * * *
# a=""
# for item in range(5):
#     a=a+"* "
#     print(a)
# 用嵌套for循环输出等腰三角形
#     *
#    * *
#   * * *
#  * * * *
# * * * * *
# a=''
# b = ""
# for item in range(5):
#     a = a + "* "
#     b = ""
#     for i in range(4 - item):
#         b = b + " "
#     print(b+a)
# for i in range(1,6):
#     print(" "*(5-i),end="")
#     print(" *"*i)
#
# s=[6,99,101,34,78,2]
# print(len(s))
# for i in range(len(s)):
#     for j in range(len(s)-1):
#        if s[j]>s[j+1]:
#            s[j],s[j+1]=s[j+1],s[j]
# print(s)

"""人和机器进行猜拳游戏写一段程序，首先选择角色：1 曹操 2张飞 3 刘备，然后选择的角色进行猜拳：1剪刀 2石头 3布 玩家输入一个1-3的数字
；然后电脑出拳 随机产生1个1-3的数字，提示电脑出拳结果（ 1剪刀 2石头 3布 ） ，
双方出拳完毕后：角色和机器出拳对战，对战结束后，最后出示本局对战结果...赢...输,
然后提示用户是否继续？按y继续，按n退出。最后结束的时候输出结果 角色赢几局 电脑赢几局，平局几次 游戏结束"""
# import random
# people=input("请选择玩家1曹操 2张飞 3 刘备：")
# s="y"
# P=0
# Y=0
# S=0
# while s=="y":
#     num=random.randint(1,3)
#     num1=input("请输入猜拳：1剪刀 2石头 3布：")
#     if num==1:
#         if num1==1:
#             print("平局")
#             P+=1
#         elif num1==2:
#             print("赢")
#             Y+=1
#         else:
#             print("输")
#             S+=1
#     if num==2:
#         if num1==1:
#             print("输")
#             Y+=1
#         elif num1==2:
#             print("平局")
#             P+=1
#         else:
#             print("赢")
#             Y+=1
#     if num==3:
#         if num1==3:
#             print("平局")
#             P+=1
#         elif num1==1:
#             print("赢")
#             Y+=1
#         else:
#             print("输")
#             S+=1
#     s=input("是否继续按y继续，按n退出：")
# print("玩家{}总共玩了{}局，赢{}局，输{}局，平{}局".format(people,Y+S+P,Y,S,P))


"""1、请编程实现字符串的转换：将”sdSdsfdAdsdsdfsfdsdASDSDFDSFa”字符串大写变小写，小写变大写。并且将字符串变为镜像字符串。
例如：字符串里面的’A’变为’Z’,’b’变为’y’ ，镜像的意思就是照镜子，对比了解下。"""
num1="AaBbCcDdEeFfGgHhIiJjKkLlMmNnOoPpQqRrSsTtVvUuWwXxYyZz"
tranttable=str.maketrans(num1,num1[::-1])
s="sdSdsfdAdsdsdfsfdsdASDSDFDSFa"
# s=s.swapcase()
print(s.translate(tranttable))

"""2：搜索引擎中会对用户输入的数据进行处理，第一步就是词法分析，分离字符串中的数字、中文、拼音、符号。
比如这个字符串： 我的是名字是lemon,今年5岁。
语法分析后得到结果如下：
数字：5
中文：我的名字是、今年、岁
拼音：lemon
符号：，。 请编写程序实现该词法分析功能。"""
# s="我的是名字是lemon,今年5岁。"
# channe=""
# numb=""
# chan=""
# strr=""
# for i in range(len(s)):
#     if s[i].isdigit()==True:
#         numb="{}{}".format(numb,s[i])
#     elif s[i].isalpha()==True:
#         if s[i]>=u"\u4e00" and s[i]<=u"\u9fa5":
#             channe="{}{}".format(channe,s[i])
#         else:
#             chan="{}{}".format(chan,s[i])
#     else:
#         strr="{}{}".format(strr,s[i])
# print("数字:{}\n中文:{}\n拼音:{}\n符号:{}".format(numb,channe,chan,strr))

"""有两行数据，存在txt文件里面：

url:http://47.107.168.87:8080/futureloan/mvc/api/member/register@mobile:18688773467@pwd:123456

url:http://47.107.168.87:8080/futureloan/mvc/api/member/recharge@mobile:18688773467@amount:1000

请利用上课所学知识，把txt里面的两行内容，写一个函数，返回如下格式的数据：

[{'url':'http://47.107.168.87:8080/futureloan/mvc/api/member/register','mobile':'18688773467','pwd':'123456'},

{'url':'http://47.107.168.87:8080/futureloan/mvc/api/member/recharge','mobile':'18688773467','amount':'1000'}]

请自行copy数据到Python里面，进行完整的分析后，再进行程序的编写！！！！

多思考多讨论！"""
# file=open("test.txt","w+",encoding="utf-8")
# file.write("url:http://47.107.168.87:8080/futureloan/mvc/api/member/register@mobile:18688773467@pwd:123456\n"
#            "url:http://47.107.168.87:8080/futureloan/mvc/api/member/recharge@mobile:18688773467@amount:1000")
# file.seek(0,0)
# res=file.readlines()
# l=[]
# for i in range(len(res)):
#    res1=res[i].strip("\n")
#    res2=res1.split("@")
#    d={}
#    for res3 in res2:
#        res4=res3.split(":",1)
#        print(res4)
#        d[res4[0]]=res4[1]
#    l.append(d)
# print(l)
# file.close()

"""1：创建一个名为 Restaurant 的类，其方法 init ()设置两个属性： restaurant_name 和 cooking_type。
创建一个名为 describe_restaurant()的方法和一个名为 open_restaurant()的方法，其中前者打印前述两项信息，
而后者打印一条消息， 指出餐馆正在营业。 根据这个类创建一个名为 restaurant 的实例，分别打印其两个属性，再调用前述两个方法。"""
# class Restaurant():
#     def __init__(self,restaurant_name,cooking_type):
#         self.restaurant_name=restaurant_name
#         self.cooking_type=cooking_type
#     def describe_restaurant(self):
#         print("餐馆名称是{}".format(self.restaurant_name))
#         print('餐馆招牌菜品{}'.format(self.cooking_type))
#
#     def open_restaurant(self):
#         print("餐厅正在营业")
#         return "餐厅正在营业"
# t=Restaurant("柒加捌杂货店","鸡肉")
# print(t.restaurant_name)
# print(t.cooking_type)
# t.describe_restaurant()
# t.open_restaurant()
"""2:建一个名为 User 的类，其中包含属性 first_name 和 last_name，还有用户简介通常会存储的其他几个属性。
在类 User 中定义一个名为 describe_user()的方法，它打印用户信息摘要；再定义一个名为 greet_user()的方法，它向用户发出个性化的问候。"""
# class User():
#     first_name="泡泡"
#     last_name="向"
#     @classmethod
#     def describe_user(cls):
#         print("你好，我是{}{}".format(cls.last_name ,cls.first_name ))
#     def greet_user(self,name,content="早上好"):
#         print("{}{}".format(name,content))
# s=User()
# s.describe_user()
# s.greet_user("wahaha")
"""3：思考问题：

#为什么会有对象方法 类方法 静态方法？我什么时候写什么类型的方法呢？
对象方法只能由对象本身调用，必须创建实体化对象才可调用
类方法在类的里面无法调用对象方法
静态方法在类里面无法调用任何方法，必须创建实体类

#什么时候用？
在类里面需要调用其方法的时候用对象方法
当类的方法独立于其他属性和方法时可以用静态方法
在类里面不需要调用对象方法时可以用类方法

#写成不同的方法类型 有啥用呢？对我来说有啥用?
调用不同，对象方法类无法调用
#他们有什么区别呢？ #自己总结一波。
@classmethod  cls  代表类方法
@staticmethod  代表静态方法
self   代表对象方法"""

"""2:写一个子类，继承1 这个父类，且添加函数：discount 打折扣用的
pay_money 支付餐费用，重写 open_restaurant() ，并完成子类完成调用"""
# class Restaurant_1(Restaurant):
#     @classmethod
#     def Price(cls,discount,pay_money):
#         print("您本次消费{}，可享受折扣{}，一共需要支付{}".format(pay_money/discount ,discount,pay_money ))
#     def open_restaurant(self,time):
#         if time>24:
#             print("时间超出范围，时间在24以内")
#         else:
#             if 11<=time and time<=24:
#                 print("餐厅正在营业")
#             else:
#                 print("餐厅停止营业，请明天再来")
# T=Restaurant_1("柒加捌杂货店","鸡肉")
# print((T.restaurant_name))
# T.Price(0.8,100)
# T.open_restaurant(13)
"""3：根据0307 0309的学习内容，来做好笔记以及思维导图"""

from class_0312 .learn_requests import Login
import unittest
from ddt import ddt,data,unpack
from openpyxl import load_workbook
wb=load_workbook("testcase.xlsx")
sheet=wb["Sheet1"]
l=[]
for i in range(1,sheet.max_row+1):
    l1=[]
    for j in range(1,sheet.max_column+1):
        resp=sheet.cell(i,j).value
        l1.append(resp)
    l.append(l1)
    wb.close()
@ddt
class Test_HttpRequest_Right(unittest.TestCase):
    T=Login()
    url="http://admin.uat.jxintell.com/user/login"
    method="post"
    @data(*l)
    @unpack
    def test_001(self,params1,expected):
        # print("001")
        # expected=200
        # params1={"phone":"15102742565","password":"123456","phoneOrPc":"PC"}
        resp=self.T.http_request(self.url,self.method,params1).status_code
        self.assertEqual(expected,resp)
#     def test_002(self):
#         print("002")
#         expected=200
#         params1={"phone":"15102000010","password":"123456","phoneOrPc":"PC"}
#         resp=self.T.http_request(self.url,self.method,params1).status_code
#         self.assertEqual(expected,resp)
# class Test_HttpRequest_Wrong(unittest.TestCase):
#     T=Login()
#     url="http://admin.uat.jxintell.com/user/login"
#     method="post"
#     def test_003(self):
#         print("003")
#         expected='{"status":500201,"message":"String index out of range: -1","data":null}'
#         params1={"phone":"","password":"456789","phoneOrPc":"PC"}
#         resp=self.T.http_request(self.url,self.method,params1).text
#         self.assertEqual(expected,resp)
#     def test_004(self):
#         print("004")
#         expected='{"status":500201,"message":"String index out of range: -1","data":null}'
#         params1={"phone":"15102742565","password":"","phoneOrPc":"PC"}
#         resp=self.T.http_request(self.url,self.method,params1).text
#         self.assertEqual(expected,resp)
# from openpyxl import load_workbook
# wb=load_workbook("login_case.xlsx")
# sheet=wb["sheet"]
# l=[]
# for i in range(1,sheet.max_row+1):
#     l1=[]
#     for j in range(1,sheet.max_cloumn+1):
#         resp=sheet.cell(i,j)
#         l1.append(resp)
#     l.append(l1)
# print(l)

