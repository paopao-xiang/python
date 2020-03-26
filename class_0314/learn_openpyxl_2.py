# -*- coding: utf-8 -*-
# @Time    : 2019/3/14 21:41
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : learn_openpyxl_2.py

#读诗  静夜思
from openpyxl import load_workbook
#
# wb=load_workbook('py15.xlsx')
#
# sheet=wb['Sheet1']
# l=[]
#
# print(sheet.max_row)#row 获取最大行
# print(sheet.max_column)#column 获取最大列
# for i in range(1,sheet.max_row+1):#控制行
#     for j in range(1,sheet.max_column+1):#1 2 3 4  为什么要加1？#控制列
#
#         if sheet.cell(i,j).value:#???
#            s=sheet.cell(i,j).value
#            print(str(s))
#            l.append(str(s))
        # print(sheet.cell(2,2).value)
        # print(sheet.cell(2,3).value)
        # print(sheet.cell(2,4).value)

#单元格为空的时候  读取到的是None


#安排一个作业
#写一个类  类里面有2个方法  1）读数据  2）写数据
from openpyxl import load_workbook
# class Load_work():
#     def __init__(self,excel_name,sheet_name,row="",column=""):#文件名称，和表单名称
#         self.excel_name=excel_name
#         self.sheet_name=sheet_name
#         self.row=row
#         self.column=column
#     def read_work(self):
#         wb=load_workbook(self.excel_name )
#         sheet=wb[self.sheet_name ]
#         if self.row !="":
#             if self.row>sheet.max_row:
#                 print("行数错误，不能超过{}".format(sheet.max_row ))
#             if self.column>sheet.max_column:
#                 print("行数错误，不能超过{}".format(sheet.max_column ))
#             print(sheet.cell(self.row,self.column).value)
#         wb.close()
#     def write_work(self,value1):
#         wb=load_workbook(self.excel_name )
#         sheet=wb[self.sheet_name ]
#         if self.row !="":
#             if self.row>sheet.max_row:
#                 print("行数错误，不能超过{}".format(sheet.max_row ))
#             if self.column>sheet.max_column:
#                 print("行数错误，不能超过{}".format(sheet.max_column ))
#             sheet.cell(self.row,self.column).value=value1
#             wb.save("py,xh.xlsx")
#             print(sheet.cell(self.row,self.column).value)
#         wb.close()
# t=Load_work('py15.xlsx','Sheet1',2,10)
# t.read_work()
# t.write_work("hahahhaha")
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，所有子列表数据放到一个大列表里面,要求把读取到的数据返回
from openpyxl import load_workbook
class Load_work():
    def __init__(self,excel_name,sheet_name):#文件名称，和表单名称
        self.excel_name=excel_name
        self.sheet_name=sheet_name
        self.wb=load_workbook(self.excel_name)
        self.sheet=self.wb[self.sheet_name]
    def read_work(self):
        L=[]
        for i in range(1,self.sheet.max_row+1):#定位最大行
            l=[]
            for j in range(1,self.sheet.max_column+1):#定位最大列
                if self.sheet.cell(i,j).value:
                    s=self.sheet.cell(i,j).value#读取每行数据=
                    l.append(s)
            L.append(l)
        self.wb.close()
        print(L)
        return L
    def write_work(self,value1,row,column):
        self.row=row
        self.column=column
        if self.row !="":
            if self.row>self.sheet.max_row:
                print("行数错误，不能超过{}".format(self.sheet.max_row ))
            if self.column>self.sheet.max_column:
                print("列数错误，不能超过{}".format(self.sheet.max_column ))
            self.sheet.cell(self.row,self.column).value=value1
            # wb.save("py,xh.xlsx")
            print(self.sheet.cell(self.row,self.column).value)
        self.wb.close()
if __name__ == '__main__':
    t=Load_work('py15.xlsx','Sheet1')
    t.read_work()
    t.write_work("hahahhaha",2,10)

#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值
#温馨提示：记得关闭和保存Excel
#
# wb=load_workbook("py15.xlsx")
# sheet=wb["Sheet1"]
# l=[]
# L=[]
# for i in range(1,sheet.max_row+1):#定位最大行
#     for j in range(1,sheet.max_column+1):#定位最大列
#         if sheet.cell(i,j).value:
#             s=sheet.cell(i,j).value#读取每行数据=
#             l.append(s)
#     L.append(l)
# print(L)
# wb.close()
