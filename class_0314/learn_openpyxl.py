# -*- coding: utf-8 -*-
# @Time    : 2019/3/14 20:27
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : learn_openpyxl.py
#openpyxl模块学习&编写类
#数据类型需要注意

#利用Python去操作Excel的话 那么就要安装 pip  install openpyxl
#创建Excel文件的模块 workbook
#读写Excel文件的模块 load_workbook

from openpyxl import workbook
# #创建一个Excel：手动  代码创建
# #新建一个Excel
wb=workbook.Workbook("cs.xlsx")
wb.create_sheet('lbb',index=1)#创建表单的方法 创建一个你自己命名的表单
wb.save("cs.xlsx")#另存为 保存工作簿
# # workbook  excel工作簿  sheet  表单  cell  表单内容

#开始操作Excel 读写
from openpyxl import  load_workbook
#读操作
#三步骤
#第一步：打开Excel 工作簿---workbook
wb=load_workbook('py15.xlsx')#打开的工作簿

#第二步：定位到表单
# sheet=wb.get_sheet_by_name('Sheet1')
sheet=wb['Sheet1']#推荐使用

#第三步：定位单元格  获取内容  根据行列坐标来定位单元格再获取值
#行列坐标都必须是数字 对应关系：  A-1 B-2 C-3 D-4 E-5
res=sheet.cell(1,2).value
print('res的值:{}，res的类型:{}'.format(res,type(res)))
# A3=sheet.cell(3,1).value
# print('res的值:{}，res的类型:{}'.format(A3,type(A3)))
# B3=sheet.cell(3,2).value
# print('res的值:{}，res的类型:{}'.format(B3,type(B3)))
# C3=eval(sheet.cell(3,3).value)
# print('res的值:{}，res的类型:{}'.format(C3,type(C3)))
# # D3=sheet.cell(3,4).value
# # print('res的值:{}，res的类型:{}'.format(D3,type(D3)))
# C4=eval(sheet.cell(4,3).value)
# print('res的值:{}，res的类型:{}'.format(C4,type(C4)))
# D4=eval(sheet.cell(4,4).value)
# print('res的值:{}，长度是：{}，res的类型:{}'.format(D4,len(D4),type(D4)))

# D4=sheet.cell(4,4).value
# print('res的值:{}，长度是：{}，res的类型:{}'.format(D4,len(D4),type(D4)))
#得出一个结论：数字还是数字 其他数据类型全是字符串类型
#eval() 可以把数据转成Python原本可以识别的数据类型
#写入值 保存

#两种方式写入值：修改和新增 都是如下2个方法
sheet.cell(5,1).value='白日依山尽'#赋值运算1
sheet.cell(3,2,'小新')#写入值2

#保存&另存为  如果是保存到当前的Excel的话 记得要关闭Excel 不然会报错 permission denied
wb.save('py15.xlsx')

#操作完毕之后，一定要关闭文件
wb.close()
# 循环读值
