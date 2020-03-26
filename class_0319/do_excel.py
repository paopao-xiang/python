# -*- coding: utf-8 -*-
# @Time    : 2019/3/19 20:05
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py
#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
# 2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值
#温馨提示：记得关闭和保存Excel
from openpyxl import load_workbook
from week_6.class_0319.read_config import ReadConfig

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#对象属性   对象可以调用  self
        self.sheet_name=sheet_name#对象属性

    def read_data(self):#Excel的文件名  一个表单名
        '''
        :file_name:目标工作簿的名称
        :sheet_name :指定的表单名'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        #从配置文件读取数据  决定获取哪些行的数据
        line=ReadConfig('case.conf').get_strValue('LINENO','line')

        all_data=[]#大列表  所有行的数据
        #读取第一行
        for i in range(1,sheet.max_row+1):#取头取尾
            sub_list=[]
            sub_list.append(sheet.cell(i,1).value)
            sub_list.append(sheet.cell(i,2).value)
            sub_list.append(sheet.cell(i,3).value)
            sub_list.append(sheet.cell(i,4).value)
            all_data.append(sub_list)

        final_data=[]#最后返回的数据
        if line=='all':#读取所有的数据
            final_data=all_data#
        else:#读取指定列表里面的指定行数的数据
            for i in eval(line):#遍历列表里面的行数的数字 eval(line) 是把字符串的line 变成列表类型的line
               final_data.append(all_data[i-1])#[1,2]
               #添加数据  line列表里面的数字 跟 all_data的数据的索引是不是-1关系
        wb.close()#关闭文件
        return final_data#返回测试数据

    def write_back(self,row,column,new_value):#写回数据的方法
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        #指定的行列值写入指定的值 new_value
        sheet.cell(row,column).value=new_value
        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
   print(DoExcel('py15.xlsx','Sheet1').read_data())




#为什么要一行存在一起的？？？cooper 用例是一行一行的 我们用的是Excel来存储测试数据